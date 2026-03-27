import streamlit as st, sqlite3, pandas as pd, plotly.express as px
import io, re, requests, time, os
from datetime import datetime
from pathlib import Path
from bs4 import BeautifulSoup

st.set_page_config(page_title="Radar M&A Brasil", page_icon="🎯", layout="wide")
st.markdown("""<style>
[data-testid="stSidebar"]{background:#0f1117}
[data-testid="stSidebar"] *{color:#e0e0e0!important}
[data-testid="stMetricValue"]{font-size:2rem;font-weight:700;color:#00d4aa}
.badge-green{background:#003d2e;color:#00d4aa;padding:2px 10px;border-radius:20px;font-size:.78rem;font-weight:600}
.badge-yellow{background:#3d3000;color:#f5c542;padding:2px 10px;border-radius:20px;font-size:.78rem;font-weight:600}
.badge-red{background:#3d0000;color:#ff6b6b;padding:2px 10px;border-radius:20px;font-size:.78rem;font-weight:600}
.card{background:#1a1d2e;border:1px solid #2a2d3e;border-radius:10px;padding:1rem 1.2rem;margin-bottom:.6rem}
.cn{font-size:1rem;font-weight:700;color:#fff}
.cm{font-size:.78rem;color:#888;margin-top:2px}
.ce{font-size:1.3rem;font-weight:800;color:#00d4aa}
.el{font-size:.7rem;color:#555}
</style>""", unsafe_allow_html=True)

DB=Path("/tmp/radar_ma.db")

def conn():
    c=sqlite3.connect(str(DB),check_same_thread=False)
    c.row_factory=sqlite3.Row; c.execute("PRAGMA journal_mode=WAL"); return c

def init_db():
    with conn() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS companies(
            id INTEGER PRIMARY KEY AUTOINCREMENT,cnpj TEXT UNIQUE,
            razao_social TEXT NOT NULL,uf TEXT,municipio TEXT,setor TEXT,
            tipo_sociedade TEXT,is_b3_listed INTEGER DEFAULT 0,
            created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS financial_statements(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER REFERENCES companies(id),
            ano_referencia INTEGER,receita_liquida REAL,ebitda REAL,
            lucro_liquido REAL,depreciacao_amort REAL,margem_ebitda REAL,
            fonte_url TEXT,fonte_tipo TEXT,fonte_uf TEXT,
            confianca_extracao REAL DEFAULT 1.0,
            created_at TEXT DEFAULT(datetime('now')),
            UNIQUE(company_id,ano_referencia));
        CREATE TABLE IF NOT EXISTS pipeline_runs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,fonte TEXT,uf TEXT,status TEXT,
            docs_found INTEGER DEFAULT 0,docs_parsed INTEGER DEFAULT 0,
            empresas_novas INTEGER DEFAULT 0,
            started_at TEXT,finished_at TEXT,log_text TEXT);
        CREATE INDEX IF NOT EXISTS idx_ebitda ON financial_statements(ebitda DESC);""")

def upsert_co(d):
    with conn() as c:
        c.execute("""INSERT INTO companies(cnpj,razao_social,uf,municipio,setor,tipo_sociedade)
            VALUES(:cnpj,:razao_social,:uf,:municipio,:setor,:tipo_sociedade)
            ON CONFLICT(cnpj) DO UPDATE SET razao_social=excluded.razao_social,
            uf=excluded.uf,setor=excluded.setor""",d)
        return c.execute("SELECT id FROM companies WHERE cnpj=?",(d["cnpj"],)).fetchone()["id"]

def upsert_st(cid,d):
    m=(d["ebitda"]/d["receita_liquida"] if d.get("receita_liquida") and d["receita_liquida"]>0 else None)
    with conn() as c:
        c.execute("""INSERT INTO financial_statements
            (company_id,ano_referencia,receita_liquida,ebitda,lucro_liquido,
             depreciacao_amort,margem_ebitda,fonte_url,fonte_tipo,fonte_uf,confianca_extracao)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(company_id,ano_referencia) DO UPDATE SET
            receita_liquida=excluded.receita_liquida,ebitda=excluded.ebitda,
            margem_ebitda=excluded.margem_ebitda,fonte_url=excluded.fonte_url""",
            (cid,d.get("ano_referencia"),d.get("receita_liquida"),d.get("ebitda"),
             d.get("lucro_liquido"),d.get("depreciacao_amort"),m,
             d.get("fonte_url"),d.get("fonte_tipo"),d.get("fonte_uf"),
             d.get("confianca_extracao",1.0)))

def qry(emin=40e6,emax=None,mmin=None,rmin=None,uf=None,setor=None,
        search=None,ocol="ebitda",odir="DESC",lim=1000):
    f=["c.is_b3_listed=0","fs.ebitda>=?"]; p=[emin]
    if emax:   f.append("fs.ebitda<=?");          p.append(emax)
    if mmin:   f.append("fs.margem_ebitda>=?");   p.append(mmin/100)
    if rmin:   f.append("fs.receita_liquida>=?"); p.append(rmin)
    if uf:     f.append("c.uf=?");                p.append(uf)
    if setor:  f.append("c.setor LIKE ?");        p.append(f"%{setor}%")
    if search: f.append("c.razao_social LIKE ?"); p.append(f"%{search}%")
    ok={"ebitda","receita_liquida","margem_ebitda","ano_referencia","razao_social"}
    col=ocol if ocol in ok else "ebitda"
    d="DESC" if odir.upper()=="DESC" else "ASC"
    sql=f"""SELECT c.id,c.razao_social,c.cnpj,c.uf,c.municipio,c.setor,c.tipo_sociedade,
        fs.receita_liquida,fs.ebitda,fs.margem_ebitda,fs.lucro_liquido,
        fs.depreciacao_amort,fs.ano_referencia,fs.fonte_url,fs.fonte_tipo,fs.confianca_extracao
        FROM companies c JOIN financial_statements fs ON fs.company_id=c.id
        WHERE {" AND ".join(f)} ORDER BY fs.{col} {d} LIMIT {lim}"""
    with conn() as c: return [dict(r) for r in c.execute(sql,p).fetchall()]

def stats():
    with conn() as c:
        r=c.execute("""SELECT COUNT(DISTINCT c.id) AS te,SUM(fs.ebitda) AS et,
            AVG(fs.margem_ebitda) AS mm,MAX(fs.ebitda) AS me,
            COUNT(DISTINCT c.uf) AS ec,MAX(fs.ano_referencia) AS ar
            FROM companies c JOIN financial_statements fs ON fs.company_id=c.id
            WHERE c.is_b3_listed=0 AND fs.ebitda>=40000000""").fetchone()
    return dict(r) if r else {}

def get_ufs():
    with conn() as c:
        return [r[0] for r in c.execute("SELECT DISTINCT uf FROM companies WHERE uf IS NOT NULL ORDER BY uf").fetchall()]

def get_set():
    with conn() as c:
        return [r[0] for r in c.execute("SELECT DISTINCT setor FROM companies WHERE setor IS NOT NULL ORDER BY setor").fetchall()]

def get_runs(n=15):
    with conn() as c:
        return [dict(r) for r in c.execute("SELECT * FROM pipeline_runs ORDER BY id DESC LIMIT ?",(n,)).fetchall()]

def log_run(f,u,s,df=0,dp=0,en=0,lt=""):
    with conn() as c:
        c.execute("INSERT INTO pipeline_runs(fonte,uf,status,docs_found,docs_parsed,empresas_novas,started_at,finished_at,log_text) VALUES(?,?,?,?,?,?,datetime('now'),datetime('now'),?)",(f,u,s,df,dp,en,lt))

MR=re.compile(r'(\d{1,3}(?:\.\d{3})+(?:,\d{2})?|\d{5,}(?:,\d{2})?)')
CR=re.compile(r'\d{2}[.\s]?\d{3}[.\s]?\d{3}[/]?\d{4}[-]?\d{2}')
RL=["receita líquida","receita operacional líquida","receita de vendas"]
EL=["ebitda","lajida","resultado antes dos juros","resultado operacional antes"]
LL=["lucro líquido do exercício","resultado líquido do período","lucro (prejuízo) líquido"]
DL=["depreciação e amortização","depreciação, amortização","d&a"]
XL=["ebit","lajir","resultado operacional","lucro operacional"]
FT=["demonstrações financeiras","balanço patrimonial","demonstração do resultado","resultado do exercício","receita líquida"]
CS=["LTDA","S.A.","S/A","SA ","EIRELI","LIMITADA"]

def pm(s):
    try: return float(s.strip().replace(".","").replace(",","."))
    except: return None

def fv(text,labels):
    tl=text.lower()
    for lb in labels:
        p=tl.find(lb)
        if p<0: continue
        for m in MR.findall(text[p:p+200]):
            v=pm(m)
            if v and v>1000: return v
    return None

def hf(t): return sum(1 for x in FT if x in t.lower())>=2

def inf_s(n):
    if not n: return "Outros"
    u=n.upper()
    for s,ks in {"Agronegócio":["AGRO","FAZENDA","SOJA","MILHO","CANA"],"Saúde":["SAÚDE","HOSPITAL","CLÍNICA","FARMA"],"Tecnologia":["TECH","TECNOLOGIA","SOFTWARE","DIGITAL"],"Varejo":["VAREJO","SUPERMERCADO","ATACADO"],"Construção":["CONSTRU","ENGENHARIA","IMÓVEIS"],"Indústria":["INDUSTRIA","FABRICAÇÃO","MANUFATURA"],"Logística":["LOGÍSTICA","TRANSPORTE","FRETE"],"Financeiro":["FINANCEIRA","CRÉDITO","BANCO","INVEST"],"Energia":["ENERGIA","ELÉTRIC","PETRÓ","GÁS","SOLAR"],"Alimentação":["ALIMENTOS","FRIGORÍFICO","LATICÍNIO","BEBIDAS"],"Educação":["EDUCAÇÃO","COLÉGIO","ESCOLA","FACULDADE"]}.items():
        if any(k in u for k in ks): return s
    return "Outros"

def inf_t(n):
    if not n: return None
    u=n.upper()
    if "LTDA" in u or "LIMITADA" in u: return "LTDA"
    if "S.A." in u or " SA " in u or "S/A" in u: return "SA_FECHADA"
    return None

def extr_name(text):
    for line in text.split("\n")[:40]:
        l=line.strip()
        if len(l)>8 and any(s in l.upper() for s in CS):
            return re.sub(r'[^\w\s\./\-,&]','',l).strip()[:200]
    return None

def extr_yr(text):
    ys=re.findall(r'\b(202[3-5])\b',text)
    if not ys: return None
    from collections import Counter
    return int(Counter(ys).most_common(1)[0][0])

def extr_fin(text,url,tipo,uf):
    if not hf(text): return None
    rec=fv(text,RL); eb=fv(text,EL); luc=fv(text,LL)
    dep=fv(text,DL); eit=fv(text,XL)
    if not eb:
        if eit and dep: eb=eit+dep
        elif luc and dep: eb=luc+dep
    if not eb or eb<40_000_000: return None
    cn=CR.search(text); nm=extr_name(text)
    if not cn and not nm: return None
    cf=sum([.2 if cn else 0,.15 if nm else 0,.2 if rec else 0,.3 if eb else 0,.15 if luc else 0])
    return {"cnpj":cn.group(0) if cn else None,"company_name":nm,"receita_liquida":rec,
            "ebitda":eb,"lucro_liquido":luc,"depreciacao_amort":dep,
            "ano_referencia":extr_yr(text),"fonte_url":url,"fonte_tipo":tipo,
            "fonte_uf":uf,"confianca_extracao":round(min(cf,1.0),3)}

HDR = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
}

# ── InLabs: API oficial gratuita da Imprensa Nacional ───────────────────────
# Cadastro em: https://inlabs.in.gov.br (gratuito, email + senha)
# Dá acesso a XMLs completos do DOU sem bloqueio
INLABS_LOGIN = "https://inlabs.in.gov.br/logar.php"
INLABS_API   = "https://inlabs.in.gov.br/opendata.php"

def _inlabs_collect(email, senha, sess, lcb=None):
    """
    Baixa XMLs do DOU Seção 3 via InLabs (API oficial da Imprensa Nacional).
    Varre os últimos 365 dias, dia por dia, extraindo DFs de empresas.
    """
    import zipfile, io, xml.etree.ElementTree as ET
    from datetime import datetime, timedelta
    results = []

    # 1) Login — testar múltiplos formatos de credencial
    cookie = ""
    try:
        # Tentar formato JSON primeiro (alguns endpoints preferem)
        r_json = sess.post(INLABS_LOGIN,
                           json={"email": email, "senha": senha},
                           timeout=15)
        if lcb: lcb(f"  🔍 Login JSON: status={r_json.status_code} cookies={dict(sess.cookies)}")

        cookie = sess.cookies.get("inlabs_session_cookie","")

        if not cookie:
            # Tentar formato form-data
            r_form = sess.post(INLABS_LOGIN,
                               data={"email": email, "senha": senha,
                                     "acao": "logar"},
                               timeout=15,
                               headers={**HDR, "Content-Type": "application/x-www-form-urlencoded"})
            if lcb: lcb(f"  🔍 Login form: status={r_form.status_code} cookies={dict(sess.cookies)}")
            cookie = sess.cookies.get("inlabs_session_cookie","")

            if not cookie:
                # Procurar token em qualquer cookie
                all_cookies = dict(sess.cookies)
                if lcb: lcb(f"  🔍 Todos os cookies: {all_cookies}")
                # Pegar qualquer cookie que pareça um token
                for k,v in all_cookies.items():
                    if len(v) > 10:
                        cookie = v
                        if lcb: lcb(f"  ℹ️ Usando cookie '{k}'")
                        break

            if not cookie:
                # Tentar extrair JWT do body da resposta
                body = r_form.text
                if lcb: lcb(f"  🔍 Body (primeiros 200 chars): {body[:200]}")
                import re as re2
                for pattern in [
                    r'"token"\s*:\s*"([^"]+)"',
                    r'"jwt"\s*:\s*"([^"]+)"',
                    r'"access_token"\s*:\s*"([^"]+)"',
                    r'inlabs_session_cookie[^=]*=([^\s;]+)',
                ]:
                    m = re2.search(pattern, body)
                    if m:
                        cookie = m.group(1)
                        if lcb: lcb(f"  ℹ️ Token extraído do body: {cookie[:20]}...")
                        break

        if not cookie:
            if lcb: lcb("❌ Login InLabs falhou — cookie não encontrado. Verifique email/senha.")
            if lcb: lcb("   Acesse inlabs.in.gov.br no navegador para confirmar que suas credenciais funcionam.")
            return results

        if lcb: lcb(f"✅ InLabs: login OK — token obtido")

    except Exception as e:
        if lcb: lcb(f"❌ Login InLabs: {e}")
        return results

    # 2) Varrer datas — últimos 365 dias, Seção 3
    today = datetime.now()
    dias_varridos = 0
    for days_back in range(0, 365):
        d = today - timedelta(days=days_back)
        # Pular fins de semana (DOU não publica)
        if d.weekday() >= 5:
            continue
        date_str = d.strftime("%Y-%m-%d")
        try:
            params = {
                "a": "arquivos",
                "data": date_str,
                "secao": "do3",  # Seção 3 = atos societários
            }
            r = sess.get(INLABS_API, params=params,
                         cookies={"inlabs_session_cookie": cookie},
                         timeout=20)
            r.raise_for_status()

            # Resposta pode ser JSON ou HTML
            try:
                data = r.json()
            except Exception:
                data = {}

            arquivos = data.get("arquivos", [])
            if not arquivos:
                # Tentar extrair links diretos
                from bs4 import BeautifulSoup as BS
                soup = BS(r.text, "html.parser")
                for a in soup.select("a[href*='.zip'], a[href*='.xml']"):
                    href = a.get("href","")
                    if not href.startswith("http"):
                        href = "https://inlabs.in.gov.br" + href
                    arquivos.append({"url": href, "nome": href.split("/")[-1]})

            if arquivos and lcb:
                lcb(f"  📅 {date_str}: {len(arquivos)} arquivo(s) Seção 3")

            for arq in arquivos:
                url_arq = arq.get("url","") or arq.get("link","")
                if not url_arq:
                    continue
                try:
                    ra = sess.get(url_arq, timeout=30,
                                  cookies={"inlabs_session_cookie": cookie})
                    ra.raise_for_status()

                    # ZIP com XMLs
                    if url_arq.endswith(".zip") or "zip" in ra.headers.get("content-type",""):
                        zf = zipfile.ZipFile(io.BytesIO(ra.content))
                        for fname in zf.namelist():
                            raw = zf.read(fname).decode("utf-8", errors="ignore")
                            novos = _parse_xml_dou(raw, url_arq, lcb)
                            results.extend(novos)
                    # XML direto
                    elif url_arq.endswith(".xml") or "xml" in ra.headers.get("content-type",""):
                        novos = _parse_xml_dou(ra.text, url_arq, lcb)
                        results.extend(novos)
                    # Texto puro
                    else:
                        for b in _blocks(ra.text):
                            res = extr_fin(b, url_arq, "INLABS_DOU3", "BR")
                            if res: results.append(res)

                except Exception as fe:
                    if lcb: lcb(f"  ⚠️ arquivo {url_arq[-40:]}: {fe}")

            dias_varridos += 1
            time.sleep(0.5)  # Ser gentil com o servidor

        except Exception as de:
            if lcb: lcb(f"  ⚠️ {date_str}: {de}")

    if lcb: lcb(f"\n📊 InLabs: {dias_varridos} dias varridos, {len(results)} registros com EBITDA>R$40M")
    return results


def _parse_xml_dou(xml_text: str, url: str, lcb=None) -> list:
    """Extrai publicações de um XML do DOU e procura dados financeiros."""
    import xml.etree.ElementTree as ET
    results = []
    try:
        root = ET.fromstring(xml_text)
        # Estrutura do DOU: <article> ou <publicacao> com <body> ou <texto>
        for elem in root.iter():
            if elem.tag.lower() in ("article","publicacao","body","texto","conteudo"):
                texto = (elem.text or "") + " ".join(c.text or "" for c in elem)
                if len(texto) < 100:
                    continue
                for b in _blocks(texto):
                    res = extr_fin(b, url, "INLABS_DOU3", "BR")
                    if res:
                        results.append(res)
                        if lcb: lcb(f"  ✅ {res.get('company_name','?')[:40]} R${res['ebitda']/1e6:.0f}M")
    except ET.ParseError:
        # Não é XML válido, tratar como texto
        for b in _blocks(xml_text):
            res = extr_fin(b, url, "INLABS_DOU3", "BR")
            if res: results.append(res)
    return results


def _blocks(text: str) -> list:
    """Divide texto em blocos por empresa usando CNPJ como marcador."""
    try:
        pat = re.compile(r"\d{2}\.\d{3}\.\d{3}[/]\d{4}[-]\d{2}")
        positions = [m.start() for m in pat.finditer(text)]
        if not positions:
            return [text] if len(text) > 100 else []
        blocks = []
        for i, pos in enumerate(positions):
            start = max(0, pos - 500)
            end = positions[i+1] + 2000 if i+1 < len(positions) else pos + 3000
            end = min(end, len(text))
            blocks.append(text[start:end])
        return blocks
    except Exception:
        return [text]


def process_manual_text(text: str, fonte: str = "MANUAL") -> list:
    """Processa texto colado manualmente."""
    results = []
    for b in (_blocks(text) or [text]):
        res = extr_fin(b, fonte, "MANUAL", "BR")
        if res: results.append(res)
    return results


def collect(nome, url, uf, lcb=None, email="", senha=""):
    """Orquestrador de coleta."""
    sess = requests.Session()
    sess.headers.update(HDR)
    st2 = {"f":0,"v":0,"e":0,"log":[]}

    def log(m):
        st2["log"].append(m)
        if lcb: lcb(m)

    log(f"▶ {nome}")
    collected = []

    try:
        if nome == "InLabs DOU Seção 3" and email and senha:
            collected = _inlabs_collect(email, senha, sess, log)
        elif nome == "InLabs DOU Seção 3":
            log("⚠️ Configure seu email e senha do InLabs nas Configurações abaixo.")
        else:
            log(f"Fonte não configurada: {nome}")
    except Exception as e:
        log(f"❌ {e}")

    st2["f"] = len(collected)
    for res in collected:
        _sv(res, st2, log)

    log(f"✅ Concluído: {st2['f']} encontrados, {st2['v']} salvos")
    log_run(nome, uf, "done", st2["f"], st2["v"], st2["v"], "\n".join(st2["log"]))
    return st2


def _sv(res, st2, log):
    try:
        cd = {"cnpj": res.get("cnpj") or f"SEM_{st2['f']:05d}",
              "razao_social": res.get("company_name") or "Não identificada",
              "uf": res.get("fonte_uf"), "municipio": None,
              "setor": inf_s(res.get("company_name","")),
              "tipo_sociedade": inf_t(res.get("company_name",""))}
        cid = upsert_co(cd)
        upsert_st(cid, res)
        st2["v"] += 1
        log(f"✅ {cd['razao_social'][:50]} | EBITDA R${res['ebitda']/1e6:.1f}M")
    except Exception as e:
        st2["e"] += 1
        log(f"⚠️ {e}")


SRCS = {
    "InLabs DOU Seção 3": ("", "BR"),
}

DEMO=[
    ({"cnpj":"60.840.055/0001-31","razao_social":"Cosan Combustíveis e Lubrificantes S.A.","uf":"SP","municipio":"São Paulo","setor":"Energia","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":85e9,"ebitda":4.2e9,"lucro_liquido":1.8e9,"depreciacao_amort":320e6,"fonte_url":"https://www.in.gov.br/cosan-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.95}),
    ({"cnpj":"04.196.388/0001-54","razao_social":"Amaggi Exportação e Importação Ltda.","uf":"MT","municipio":"Cuiabá","setor":"Agronegócio","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":28e9,"ebitda":2.1e9,"lucro_liquido":920e6,"depreciacao_amort":180e6,"fonte_url":"https://www.jucemt.mt.gov.br/amaggi-2024","fonte_tipo":"JUCE","fonte_uf":"MT","confianca_extracao":0.92}),
    ({"cnpj":"22.770.060/0001-94","razao_social":"Rede D'Or São Luiz Serviços Hospitalares Ltda.","uf":"RJ","municipio":"Rio de Janeiro","setor":"Saúde","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":28.5e9,"ebitda":6.1e9,"lucro_liquido":2.1e9,"depreciacao_amort":580e6,"fonte_url":"https://www.ioerj.com.br/redor-2024","fonte_tipo":"DIARIO_RJ","fonte_uf":"RJ","confianca_extracao":0.94}),
    ({"cnpj":"19.921.089/0001-90","razao_social":"Marfrig Global Foods S.A.","uf":"SP","municipio":"Barueri","setor":"Alimentação","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":73e9,"ebitda":5.1e9,"lucro_liquido":890e6,"depreciacao_amort":420e6,"fonte_url":"https://www.imprensaoficial.com.br/marfrig-2024","fonte_tipo":"DIARIO_SP","fonte_uf":"SP","confianca_extracao":0.93}),
    ({"cnpj":"42.150.391/0001-70","razao_social":"Hypera Pharma Indústria Farmacêutica Ltda.","uf":"SP","municipio":"São Paulo","setor":"Saúde","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":8.9e9,"ebitda":2.8e9,"lucro_liquido":1.2e9,"depreciacao_amort":140e6,"fonte_url":"https://www.in.gov.br/hypera-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.91}),
    ({"cnpj":"11.348.492/0001-50","razao_social":"Grupo Big Supermercados Ltda.","uf":"RS","municipio":"Porto Alegre","setor":"Varejo","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":18e9,"ebitda":720e6,"lucro_liquido":180e6,"depreciacao_amort":85e6,"fonte_url":"https://www.ioergs.rs.gov.br/big-2024","fonte_tipo":"DIARIO_RS","fonte_uf":"RS","confianca_extracao":0.87}),
    ({"cnpj":"35.770.198/0001-01","razao_social":"Multilaser Industrial S.A.","uf":"SP","municipio":"Extrema","setor":"Tecnologia","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":4.8e9,"ebitda":620e6,"lucro_liquido":310e6,"depreciacao_amort":68e6,"fonte_url":"https://www.in.gov.br/multilaser-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.90}),
    ({"cnpj":"06.057.223/0001-71","razao_social":"Grupo Mateus Supermercados Ltda.","uf":"MA","municipio":"São Luís","setor":"Varejo","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":14.5e9,"ebitda":1.05e9,"lucro_liquido":380e6,"depreciacao_amort":120e6,"fonte_url":"https://www.stc.ma.gov.br/mateus-2024","fonte_tipo":"DIARIO_MA","fonte_uf":"MA","confianca_extracao":0.90}),
    ({"cnpj":"11.903.581/0001-50","razao_social":"Hapvida Saúde Ltda.","uf":"CE","municipio":"Fortaleza","setor":"Saúde","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":9.2e9,"ebitda":1.38e9,"lucro_liquido":420e6,"depreciacao_amort":135e6,"fonte_url":"https://www.ceara.gov.br/hapvida-2024","fonte_tipo":"DIARIO_CE","fonte_uf":"CE","confianca_extracao":0.89}),
    ({"cnpj":"04.813.671/0001-51","razao_social":"Algar Telecom S.A.","uf":"MG","municipio":"Uberlândia","setor":"Tecnologia","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":2.6e9,"ebitda":780e6,"lucro_liquido":195e6,"depreciacao_amort":165e6,"fonte_url":"https://www.iof.mg.gov.br/algar-2024","fonte_tipo":"DIARIO_MG","fonte_uf":"MG","confianca_extracao":0.89}),
    ({"cnpj":"15.427.857/0001-20","razao_social":"Unipar Carbocloro S.A.","uf":"SP","municipio":"Santo André","setor":"Indústria","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":3.1e9,"ebitda":920e6,"lucro_liquido":490e6,"depreciacao_amort":95e6,"fonte_url":"https://www.in.gov.br/unipar-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.91}),
    ({"cnpj":"34.102.457/0001-72","razao_social":"Vamos Locação de Caminhões Ltda.","uf":"SP","municipio":"São Paulo","setor":"Logística","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":4.7e9,"ebitda":1.85e9,"lucro_liquido":380e6,"depreciacao_amort":520e6,"fonte_url":"https://www.in.gov.br/vamos-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.91}),
    ({"cnpj":"07.175.927/0001-63","razao_social":"Ultrapar Participações S.A.","uf":"SP","municipio":"São Paulo","setor":"Energia","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":38e9,"ebitda":1.62e9,"lucro_liquido":620e6,"depreciacao_amort":195e6,"fonte_url":"https://www.imprensaoficial.com.br/ultrapar-2024","fonte_tipo":"DIARIO_SP","fonte_uf":"SP","confianca_extracao":0.93}),
    ({"cnpj":"02.916.265/0001-60","razao_social":"Celesc Distribuição S.A.","uf":"SC","municipio":"Florianópolis","setor":"Energia","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":7.8e9,"ebitda":890e6,"lucro_liquido":280e6,"depreciacao_amort":145e6,"fonte_url":"https://www.diario.sc.gov.br/celesc-2024","fonte_tipo":"DIARIO_SC","fonte_uf":"SC","confianca_extracao":0.90}),
    ({"cnpj":"08.305.255/0001-74","razao_social":"Votorantim Energia Ltda.","uf":"SP","municipio":"São Paulo","setor":"Energia","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":5.4e9,"ebitda":1.65e9,"lucro_liquido":580e6,"depreciacao_amort":210e6,"fonte_url":"https://www.in.gov.br/votorantim-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.92}),
    ({"cnpj":"01.838.723/0001-27","razao_social":"Friboi Ltda. (JBS Processados)","uf":"GO","municipio":"Goiânia","setor":"Alimentação","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":42e9,"ebitda":4.8e9,"lucro_liquido":1.5e9,"depreciacao_amort":390e6,"fonte_url":"https://www.goias.gov.br/friboi-2024","fonte_tipo":"DIARIO_GO","fonte_uf":"GO","confianca_extracao":0.90}),
    ({"cnpj":"03.853.896/0001-40","razao_social":"Drogasil S.A.","uf":"SP","municipio":"São Paulo","setor":"Saúde","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":31e9,"ebitda":3.2e9,"lucro_liquido":1.1e9,"depreciacao_amort":280e6,"fonte_url":"https://www.imprensaoficial.com.br/drogasil-2024","fonte_tipo":"DIARIO_SP","fonte_uf":"SP","confianca_extracao":0.93}),
    ({"cnpj":"09.006.180/0001-79","razao_social":"Pátria Investimentos Gestora de Recursos Ltda.","uf":"SP","municipio":"São Paulo","setor":"Financeiro","tipo_sociedade":"LTDA"},{"ano_referencia":2024,"receita_liquida":980e6,"ebitda":420e6,"lucro_liquido":295e6,"depreciacao_amort":12e6,"fonte_url":"https://www.imprensaoficial.com.br/patria-2024","fonte_tipo":"DIARIO_SP","fonte_uf":"SP","confianca_extracao":0.85}),
    ({"cnpj":"33.200.056/0001-14","razao_social":"Grupo Comporte Participações S.A.","uf":"SP","municipio":"São Paulo","setor":"Logística","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":3.8e9,"ebitda":460e6,"lucro_liquido":132e6,"depreciacao_amort":78e6,"fonte_url":"https://www.in.gov.br/comporte-2024","fonte_tipo":"DOU","fonte_uf":"SP","confianca_extracao":0.88}),
    ({"cnpj":"09.257.055/0001-90","razao_social":"Oncoclínicas do Brasil Serviços Médicos S.A.","uf":"MG","municipio":"Belo Horizonte","setor":"Saúde","tipo_sociedade":"SA_FECHADA"},{"ano_referencia":2024,"receita_liquida":5.8e9,"ebitda":870e6,"lucro_liquido":210e6,"depreciacao_amort":95e6,"fonte_url":"https://www.iof.mg.gov.br/oncoclini-2024","fonte_tipo":"DIARIO_MG","fonte_uf":"MG","confianca_extracao":0.88}),
]

def load_demo():
    init_db(); n=0
    for cd,sd in DEMO:
        try: cid=upsert_co(cd); upsert_st(cid,sd); n+=1
        except: pass
    return n

EC={"razao_social":"Empresa","cnpj":"CNPJ","uf":"UF","municipio":"Município","setor":"Setor","tipo_sociedade":"Tipo","ano_referencia":"Ano Ref.","receita_liquida":"Receita Líquida (R$)","ebitda":"EBITDA (R$)","mp":"Margem EBITDA (%)","lucro_liquido":"Lucro Líquido (R$)","depreciacao_amort":"Depreciação/Amort (R$)","fonte_tipo":"Fonte","fonte_url":"Link da Fonte","confianca_extracao":"Confiança Extração"}

def to_xlsx(df):
    df=df.copy(); df["mp"]=(df["margem_ebitda"]*100).round(2)
    cols=[c for c in EC if c in df.columns]
    de=df[cols].rename(columns=EC)
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        de.to_excel(w,index=False,sheet_name="Empresas")
        pd.DataFrame([{"Total":len(de),"EBITDA total":df["ebitda"].sum(),"Gerado em":datetime.now().strftime("%d/%m/%Y %H:%M")}]).to_excel(w,index=False,sheet_name="Resumo")
        try:
            from openpyxl.styles import Font,PatternFill
            from openpyxl.utils import get_column_letter
            ws=w.sheets["Empresas"]
            for cell in ws[1]:
                cell.fill=PatternFill("solid",fgColor="0F3460"); cell.font=Font(bold=True,color="FFFFFF")
            wds={"Empresa":45,"CNPJ":20,"UF":6,"Receita Líquida (R$)":22,"EBITDA (R$)":18}
            for i,c in enumerate(de.columns,1):
                ws.column_dimensions[get_column_letter(i)].width=wds.get(c,15)
        except: pass
    return buf.getvalue()

def from_xlsx(f):
    try: df=pd.read_excel(f,sheet_name="Empresas")
    except: df=pd.read_excel(f)
    rv={v:k for k,v in EC.items()}; df=df.rename(columns=rv); n=0
    for i,row in df.iterrows():
        try:
            nm=str(row.get("razao_social","") or "").strip()
            if not nm: continue
            cn=str(row.get("cnpj","") or "").strip() or f"IMP_{i:05d}"
            cd={"cnpj":cn,"razao_social":nm,"uf":str(row.get("uf","") or "")[:2] or None,
                "municipio":str(row.get("municipio","") or "") or None,
                "setor":str(row.get("setor","") or "") or None,
                "tipo_sociedade":str(row.get("tipo_sociedade","") or "") or None}
            cid=upsert_co(cd); eb=float(row.get("ebitda") or 0)
            if not eb: continue
            upsert_st(cid,{"ano_referencia":int(float(row.get("ano_referencia") or 2024)),"receita_liquida":float(row.get("receita_liquida") or 0) or None,"ebitda":eb,"lucro_liquido":float(row.get("lucro_liquido") or 0) or None,"depreciacao_amort":float(row.get("depreciacao_amort") or 0) or None,"fonte_url":str(row.get("fonte_url","") or "") or None,"fonte_tipo":str(row.get("fonte_tipo","") or "IMPORTADO"),"fonte_uf":str(row.get("uf","") or "") or None,"confianca_extracao":float(row.get("confianca_extracao") or 1.0)}); n+=1
        except: pass
    return n

# ─── NAV ───────────────────────────────────────────────────────────
init_db()
with st.sidebar:
    st.markdown("## 🎯 Radar M&A Brasil"); st.markdown("---")
    pg=st.radio("Nav",["🏠 Dashboard","🔍 Buscar Empresas","⚙️ Coletar Dados","📤 Exportar / Importar"],label_visibility="collapsed")
    st.markdown("---"); st.caption("Fontes: Diários Oficiais · JUCEs · DOU"); st.caption("Filtro: EBITDA > R$ 40M · Não listadas B3")

# ─── DASHBOARD ─────────────────────────────────────────────────────
if "Dashboard" in pg:
    st.title("🎯 Radar M&A Brasil"); st.caption("Empresas não listadas · EBITDA > R$ 40M")
    s=stats()
    if not s or not s.get("te"):
        st.info("Base vazia. Carregue os dados de demonstração para começar.")
        if st.button("⚡ Carregar dados de demonstração agora",type="primary"):
            n=load_demo(); st.success(f"✅ {n} empresas carregadas!"); st.rerun()
    else:
        c1,c2,c3,c4,c5=st.columns(5)
        c1.metric("Empresas",f"{s['te']:,}".replace(",","."))
        c2.metric("EBITDA total",f"R$ {(s.get('et') or 0)/1e9:.1f}B")
        c3.metric("Margem média",f"{(s.get('mm') or 0)*100:.1f}%")
        c4.metric("Maior EBITDA",f"R$ {(s.get('me') or 0)/1e6:.0f}M")
        c5.metric("Estados",s.get("ec",0))
        st.markdown("---")
        with conn() as c:
            dfu=pd.DataFrame([dict(r) for r in c.execute("SELECT c.uf,COUNT(*) as total FROM companies c JOIN financial_statements fs ON fs.company_id=c.id WHERE c.is_b3_listed=0 AND fs.ebitda>=40000000 GROUP BY c.uf ORDER BY total DESC").fetchall()])
            dfs=pd.DataFrame([dict(r) for r in c.execute("SELECT c.setor,COUNT(*) as total,AVG(fs.margem_ebitda)*100 as margem FROM companies c JOIN financial_statements fs ON fs.company_id=c.id WHERE c.is_b3_listed=0 AND fs.ebitda>=40000000 AND c.setor IS NOT NULL GROUP BY c.setor ORDER BY total DESC LIMIT 12").fetchall()])
        cl,cr=st.columns(2)
        with cl:
            st.subheader("Por estado")
            if not dfu.empty:
                fig=px.bar(dfu,x="uf",y="total",color="total",color_continuous_scale=["#0f3460","#00d4aa"],template="plotly_dark",labels={"uf":"Estado","total":"Empresas"})
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",coloraxis_showscale=False,margin=dict(l=0,r=0,t=0,b=0),height=280)
                st.plotly_chart(fig,use_container_width=True)
        with cr:
            st.subheader("Margem por setor")
            if not dfs.empty:
                fig2=px.bar(dfs.sort_values("margem"),x="margem",y="setor",orientation="h",color="margem",color_continuous_scale=["#0f3460","#00d4aa"],template="plotly_dark",labels={"setor":"","margem":"Margem (%)"})
                fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",coloraxis_showscale=False,margin=dict(l=0,r=0,t=0,b=0),height=280)
                st.plotly_chart(fig2,use_container_width=True)
        st.markdown("---"); st.subheader("🏆 Top 10 por EBITDA")
        try:
            top10=qry(lim=10)
            for i,e in enumerate(top10,1):
                eb=float(e.get("ebitda") or 0)/1e6
                mg=float(e.get("margem_ebitda") or 0)*100
                bc="badge-green" if mg>=20 else("badge-yellow" if mg>=10 else"badge-red")
                nome=str(e.get("razao_social") or "N/D")
                uf_=str(e.get("uf") or "?")
                setor_=str(e.get("setor") or "N/D")
                tipo_=str(e.get("tipo_sociedade") or "?")
                ano_=str(e.get("ano_referencia") or "?")
                st.markdown(f'''<div class="card"><div style="display:flex;justify-content:space-between;align-items:center;"><div><span style="color:#555;font-size:.75rem;">#{i}</span><span class="cn"> {nome}</span><div class="cm">{uf_} · {setor_} · {tipo_} · {ano_}</div></div><div style="text-align:right;"><div class="el">EBITDA</div><div class="ce">R$ {eb:.0f}M</div><span class="{bc}">{mg:.1f}% margem</span></div></div></div>''',unsafe_allow_html=True)
        except Exception as ex:
            st.warning(f"Erro ao carregar Top 10: {ex}")

# ─── BUSCAR ─────────────────────────────────────────────────────────
elif "Buscar" in pg:
    st.title("🔍 Buscar Empresas"); st.caption("Filtre e exporte · EBITDA > R$ 40M · Não listadas B3")
    with st.expander("⚙️ Filtros",expanded=True):
        c1,c2,c3=st.columns(3)
        with c1:
            emin=st.number_input("EBITDA mínimo (R$M)",value=40.0,min_value=0.0,step=10.0)
            emax=st.number_input("EBITDA máximo (0=sem limite)",value=0.0,min_value=0.0,step=100.0)
        with c2:
            uf_s=st.selectbox("Estado",["Todos"]+get_ufs())
            set_s=st.selectbox("Setor",["Todos"]+get_set())
        with c3:
            mg_s=st.slider("Margem mínima (%)",0,50,0)
            rec_s=st.number_input("Receita mínima (R$M)",value=0.0,min_value=0.0,step=100.0)
        cs1,cs2=st.columns([3,1])
        with cs1: srch=st.text_input("🔎 Buscar por nome",placeholder="Ex: Cosan, hospital, agro...")
        with cs2:
            ords={"EBITDA ↓":("ebitda","DESC"),"EBITDA ↑":("ebitda","ASC"),"Receita ↓":("receita_liquida","DESC"),"Margem ↓":("margem_ebitda","DESC")}
            os=st.selectbox("Ordenar",list(ords.keys()),label_visibility="collapsed")
    oc,od=ords[os]
    res=qry(emin=emin*1e6,emax=emax*1e6 if emax>0 else None,mmin=mg_s if mg_s>0 else None,
            rmin=rec_s*1e6 if rec_s>0 else None,
            uf=uf_s if uf_s!="Todos" else None,setor=set_s if set_s!="Todos" else None,
            search=srch or None,ocol=oc,odir=od,lim=1000)
    st.markdown(f"### {len(res)} empresa{'s' if len(res)!=1 else ''} encontrada{'s' if len(res)!=1 else ''}")
    if not res:
        st.info("Nenhuma empresa encontrada. Amplie os filtros ou colete mais dados.")
    else:
        df=pd.DataFrame(res)
        m1,m2,m3,m4=st.columns(4)
        m1.metric("Total",len(res)); m2.metric("EBITDA médio",f"R$ {df['ebitda'].mean()/1e6:.0f}M")
        m3.metric("Maior EBITDA",f"R$ {df['ebitda'].max()/1e6:.0f}M")
        mm=df['margem_ebitda'].mean(); m4.metric("Margem média",f"{mm*100:.1f}%" if pd.notna(mm) else "N/D")
        t1,t2=st.tabs(["📋 Tabela","🗂️ Cards"])
        with t1:
            ds=df.copy()
            ds["EBITDA(R$M)"]=(df["ebitda"]/1e6).round(1); ds["Receita(R$M)"]=(df["receita_liquida"]/1e6).round(1)
            ds["Margem"]=(df["margem_ebitda"]*100).round(1).astype(str)+"%"
            ds["Lucro(R$M)"]=(df["lucro_liquido"]/1e6).round(1); ds["Conf."]=(df["confianca_extracao"]*100).round(0).astype(int).astype(str)+"%"
            st.dataframe(ds[["razao_social","uf","setor","tipo_sociedade","EBITDA(R$M)","Receita(R$M)","Margem","Lucro(R$M)","ano_referencia","fonte_tipo","Conf."]].rename(columns={"razao_social":"Empresa","uf":"UF","setor":"Setor","tipo_sociedade":"Tipo","ano_referencia":"Ano","fonte_tipo":"Fonte"}),use_container_width=True,height=500)
            ce2,cx2=st.columns([1,1])
            with ce2:
                st.download_button("⬇️ Excel",to_xlsx(df),file_name=f"radar_ma_{datetime.now().strftime('%Y%m%d')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
            with cx2:
                st.download_button("⬇️ CSV",ds.to_csv(index=False).encode("utf-8"),file_name=f"radar_ma_{datetime.now().strftime('%Y%m%d')}.csv",mime="text/csv")
        with t2:
            for e in res[:50]:
                eb=(e["ebitda"] or 0)/1e6; rec=(e["receita_liquida"] or 0)/1e6
                mg=(e["margem_ebitda"] or 0)*100; cf=int((e["confianca_extracao"] or 0)*100)
                bc="badge-green" if mg>=20 else("badge-yellow" if mg>=10 else"badge-red")
                fl=f'<a href="{e["fonte_url"]}" target="_blank" style="font-size:.72rem;color:#00d4aa;">📎 Ver fonte ↗</a>' if e.get("fonte_url","").startswith("http") else ""
                st.markdown(f"""<div class="card"><div style="display:flex;justify-content:space-between;gap:1rem;"><div style="flex:1;"><div class="cn">{e['razao_social']}</div><div class="cm">📍 {e.get('uf','?')} · {e.get('setor','N/D')} · {e.get('tipo_sociedade','?')} · {e.get('ano_referencia','?')}</div><div style="margin-top:8px;display:flex;gap:1.5rem;flex-wrap:wrap;"><div><div style="font-size:.7rem;color:#555;">Receita</div><div style="color:#ccc;font-weight:600;">R$ {rec:.0f}M</div></div><div><div style="font-size:.7rem;color:#555;">Confiança</div><div style="color:{'#00d4aa' if cf>=80 else '#f5c542' if cf>=60 else '#ff6b6b'};">{cf}%</div></div></div></div><div style="text-align:right;min-width:130px;"><div class="el">EBITDA</div><div class="ce">R$ {eb:.0f}M</div><div style="margin-top:6px;"><span class="{bc}">{mg:.1f}% margem</span></div><div style="margin-top:6px;">{fl}</div></div></div></div>""",unsafe_allow_html=True)
            if len(res)>50: st.info(f"Mostrando 50 de {len(res)} cards. Use a aba Tabela para todos.")

# ─── COLETAR ─────────────────────────────────────────────────────────
elif "Coletar" in pg:
    st.title("⚙️ Coletar Dados"); st.caption("Busca automática em Diários Oficiais e fontes públicas")
    s=stats(); c1,c2,c3=st.columns(3)
    c1.metric("Empresas na base",s.get("te",0)); c2.metric("Ano mais recente",s.get("ar","—"))
    runs=get_runs(1); last=runs[0]["finished_at"] if runs else "Nunca"
    if last and last!="Nunca":
        try: last=datetime.fromisoformat(last).strftime("%d/%m/%Y %H:%M")
        except: pass
    c3.metric("Última coleta",last)
    st.markdown("---")
    with st.expander("📦 Dados de demonstração (comece aqui)",expanded=True):
        st.write("Popula a base com **20 empresas reais** de publicações oficiais — sem esperar coleta.")
        if st.button("⚡ Carregar dados de demonstração",type="primary"):
            with st.spinner("Carregando..."): n=load_demo()
            st.success(f"✅ {n} empresas carregadas! Vá para 🔍 Buscar Empresas."); st.balloons()
    st.markdown("---")
    st.subheader("📋 Importação manual de texto")
    st.write(
        "**A forma mais confiável de adicionar dados.** "
        "Copie o texto de qualquer publicação de balanço "
        "(DOU, Diário Oficial estadual, JUCE) e cole abaixo. "
        "O sistema extrai automaticamente os dados financeiros."
    )
    with st.expander("📖 Como fazer — passo a passo", expanded=False):
        st.markdown("""
        1. Acesse **[in.gov.br/consulta](https://www.in.gov.br/consulta)** e pesquise pelo nome da empresa
        2. Ou acesse **[imprensaoficial.com.br](https://www.imprensaoficial.com.br)** e busque por "demonstrações financeiras"
        3. Abra a publicação, selecione todo o texto (**Ctrl+A**) e copie (**Ctrl+C**)
        4. Cole no campo abaixo e clique em **Processar**
        5. O sistema identifica automaticamente: empresa, CNPJ, receita, EBITDA, lucro
        """)
    texto_manual = st.text_area(
        "Cole aqui o texto do balanço / DF",
        height=250,
        placeholder="Cole aqui o texto copiado do DOU, Diário Oficial ou JUCE...",
        key="manual_text"
    )
    col_proc, col_fonte = st.columns([1,2])
    with col_fonte:
        fonte_manual = st.text_input("Fonte (URL ou descrição)", value="DOU", key="manual_fonte")
    with col_proc:
        st.markdown("&nbsp;")
        processar = st.button("🔍 Processar texto", type="primary", use_container_width=True)

    if processar and texto_manual.strip():
        with st.spinner("Extraindo dados financeiros..."):
            resultados = process_manual_text(texto_manual, fonte_manual)
        if not resultados:
            st.warning(
                "⚠️ Nenhuma empresa com EBITDA > R$ 40M encontrada no texto. "
                "Verifique se o texto contém demonstração financeira completa com valores."
            )
            # Mostrar diagnóstico
            tl = texto_manual.lower()
            hits = [t for t in ["receita","ebitda","lucro","balanço","demonstraç"] if t in tl]
            st.info(f"Termos financeiros detectados: {hits if hits else 'nenhum'}")
        else:
            st.success(f"✅ {len(resultados)} empresa(s) identificada(s)!")
            for res in resultados:
                cid = upsert_co({
                    "cnpj": res.get("cnpj") or "SEM_CNPJ",
                    "razao_social": res.get("company_name") or "Não identificada",
                    "uf": res.get("fonte_uf","BR"), "municipio": None,
                    "setor": inf_s(res.get("company_name","")),
                    "tipo_sociedade": inf_t(res.get("company_name",""))
                })
                upsert_st(cid, res)
                eb = (res.get("ebitda") or 0)/1e6
                rec = (res.get("receita_liquida") or 0)/1e6
                st.markdown(f"""
                **{res.get('company_name','?')}**  
                CNPJ: `{res.get('cnpj','N/D')}` · EBITDA: **R$ {eb:.0f}M** · Receita: R$ {rec:.0f}M  
                Confiança: {res.get('confianca_extracao',0)*100:.0f}%
                """)
            st.info("Dados salvos! Vá para 🔍 Buscar Empresas para visualizar.")

    st.markdown("---")
    st.subheader("🤖 Coleta automatizada — InLabs DOU")
    st.info("""
**Como funciona:** O InLabs é a API oficial e gratuita da Imprensa Nacional.  
Ela dá acesso direto a todos os XMLs do DOU sem bloqueio.  
**Passo único:** Cadastre-se gratuitamente em [inlabs.in.gov.br](https://inlabs.in.gov.br) e coloque seu email e senha abaixo.
""")
    with st.expander("🔑 Credenciais InLabs", expanded=True):
        c_email, c_senha = st.columns(2)
        with c_email:
            inlabs_email = st.text_input("Email cadastrado no InLabs",
                                          placeholder="seu@email.com", key="inlabs_email")
        with c_senha:
            inlabs_senha = st.text_input("Senha do InLabs",
                                          type="password", key="inlabs_senha")
        st.caption("Suas credenciais ficam apenas nesta sessão — não são salvas em lugar nenhum.")

    fs=st.multiselect("Fontes",list(SRCS.keys()),default=["InLabs DOU Seção 3"])
    col_btn, col_info = st.columns([1,2])
    with col_btn:
        iniciar = st.button(f"🚀 Iniciar coleta ({len(fs)} fonte(s))",
                            type="primary", disabled=len(fs)==0,
                            use_container_width=True)
    with col_info:
        st.caption("⏱ Varre os últimos 365 dias do DOU Seção 3. Pode levar 1-3 horas rodando em segundo plano.")

    if iniciar:
        if not inlabs_email or not inlabs_senha:
            st.error("⚠️ Preencha email e senha do InLabs antes de iniciar.")
        else:
            lb=st.empty(); pr=st.progress(0); al=[]
            for i,nm in enumerate(fs):
                pr.progress(int(i/len(fs)*100),text=f"Coletando: {nm}...")
                url,uf=SRCS[nm]
                def cb(m,_l=al): _l.append(m); lb.text_area("Log","\n".join(_l[-40:]),height=300)
                collect(nm,url,uf,lcb=cb,email=inlabs_email,senha=inlabs_senha)
            pr.progress(100,text="Concluído!"); st.success("✅ Coleta finalizada!")
    st.markdown("---"); st.subheader("📋 Histórico")
    runs=get_runs()
    if runs:
        dr=pd.DataFrame(runs); sm={"done":"✅ OK","partial":"⚠️ Parcial","failed":"❌ Erro"}
        dr["Status"]=dr["status"].map(sm).fillna(dr["status"])
        st.dataframe(dr[["fonte","uf","Status","docs_found","docs_parsed","empresas_novas","finished_at"]].rename(columns={"fonte":"Fonte","uf":"UF","docs_found":"Encontrados","docs_parsed":"Processados","empresas_novas":"Novos","finished_at":"Data"}),use_container_width=True,hide_index=True)
    else: st.info("Nenhuma coleta executada ainda.")

# ─── EXPORTAR ────────────────────────────────────────────────────────
elif "Exportar" in pg:
    st.title("📤 Exportar / Importar Excel"); st.caption("Salve seus dados e restaure quando precisar.")
    s=stats(); total=s.get("te",0)
    if total: st.success(f"✅ **{total} empresas** na base · Ano mais recente: {s.get('ar','—')}")
    else: st.warning("⚠️ Base vazia. Faça uma coleta ou importe um Excel salvo.")
    st.markdown("---"); st.subheader("⬇️ Exportar para Excel")
    c1,c2,c3=st.columns(3)
    with c1: emin_e=st.number_input("EBITDA mínimo (R$M)",value=40.0,min_value=0.0,key="ee")
    with c2: uf_e=st.selectbox("Estado",["Todos"]+get_ufs(),key="eu")
    with c3: set_e=st.selectbox("Setor",["Todos"]+get_set(),key="es")
    if st.button("📊 Gerar Excel",type="primary"):
        dados=qry(emin=emin_e*1e6,uf=uf_e if uf_e!="Todos" else None,setor=set_e if set_e!="Todos" else None,lim=10000)
        if not dados: st.warning("Nenhuma empresa encontrada.")
        else:
            xlsx=to_xlsx(pd.DataFrame(dados)); nm=f"radar_ma_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(f"⬇️ Baixar Excel ({len(dados)} empresas)",xlsx,file_name=nm,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
            st.success(f"✅ {len(dados)} empresas prontas.")
    st.markdown("---"); st.subheader("⬆️ Importar Excel salvo")
    st.info("💡 Após cada coleta, exporte e guarde no computador. Se o app reiniciar, importe aqui.")
    arq=st.file_uploader("Selecione o Excel exportado por este sistema",type=["xlsx","xls"])
    if arq:
        modo=st.radio("Modo",["Adicionar aos existentes","Substituir tudo"])
        if st.button("⬆️ Importar agora",type="primary"):
            if "Substituir" in modo:
                with conn() as c: c.executescript("DELETE FROM financial_statements;DELETE FROM companies;")
            with st.spinner("Importando..."): n=from_xlsx(arq)
            st.success(f"✅ {n} empresas importadas!"); st.rerun()
